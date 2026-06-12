# =========================
# INTELLIGENCE ROUTES FINANCE
# =========================

# =========================
# IMPORTS
# =========================
import csv
import io
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import text
from database import engine
from auth.utils import get_current_user
from core.cache import redis_client
from intelligence.gamification.progress_service import award_xp

router = APIRouter(tags=["Finance"])

# =========================
# HELPER
# =========================
def get_user_id(conn, email: str):
    row = conn.execute(
        text("""
            SELECT id
            FROM users
            WHERE email = :email
        """),
        {"email": email}
    ).fetchone()

    return row.id if row else None

# =========================
# INVALIDATE FINANCE CACHE
# =========================
def invalidate_finance_caches(email: str, user_id: Optional[int] = None):
    try:
        if not redis_client:
            return

        keys = [
            f"intel:{email}",
            f"context:{email}",
            f"score:{email}",
        ]

        if user_id is not None:
            keys.append(f"financial:{user_id}")
            keys.append(f"financial_overview:{user_id}")

        redis_client.delete(*keys)
    except Exception:
        pass


def get_item_name(data: dict):
    return data.get("name") or data.get("label") or ""


def parse_amount(value):
    cleaned = str(value or "0").strip().replace("\u202f", "").replace(" ", "")
    cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def normalize_finance_type(value: str, scope: str):
    raw = str(value or "").strip().lower()
    aliases = {
        "income": "revenus",
        "revenue": "revenus",
        "revenu": "revenus",
        "revenus": "revenus",
        "expense": "charges",
        "charge": "charges",
        "charges": "charges",
        "saving": "epargne",
        "savings": "epargne",
        "epargne": "epargne",
        "debt": "dettes",
        "debts": "dettes",
        "dette": "dettes",
        "dettes": "dettes",
    }
    item_type = aliases.get(raw)
    allowed = {"cashflow": {"revenus", "charges"}, "balance": {"epargne", "dettes"}}

    if item_type in allowed.get(scope, set()):
        return item_type

    return None


def normalize_import_row(row: dict):
    return {
        str(key or "").strip().lower(): str(value or "").strip()
        for key, value in row.items()
    }


def read_csv_rows(raw: bytes):
    try:
        content = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        content = raw.decode("latin-1")

    reader = csv.DictReader(io.StringIO(content))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV vide ou colonnes manquantes")

    return [normalize_import_row(row) for row in reader]


def read_excel_rows(raw: bytes):
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Fichier Excel invalide ou illisible")

    rows = workbook.active.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        raise HTTPException(status_code=400, detail="Excel vide ou colonnes manquantes")

    normalized_headers = [str(header or "").strip().lower() for header in headers]
    if not any(normalized_headers):
        raise HTTPException(status_code=400, detail="Excel vide ou colonnes manquantes")

    return [
        normalize_import_row(dict(zip(normalized_headers, values)))
        for values in rows
        if any(value not in (None, "") for value in values)
    ]


def read_import_rows(filename: str, content_type: str | None, raw: bytes):
    is_excel = filename.endswith(".xlsx") or content_type == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if is_excel:
        return read_excel_rows(raw), "excel"

    if filename.endswith(".csv") or content_type in {
        "text/csv",
        "application/vnd.ms-excel",
        "application/csv",
    }:
        return read_csv_rows(raw), "csv"

    raise HTTPException(status_code=400, detail="Format supporte: CSV ou Excel (.xlsx)")


# =========================
# CREATE ITEM
# =========================
@router.post("")
@router.post("/")
def create_finance_item(data: dict, user=Depends(get_current_user)):

    email = user

    with engine.begin() as conn:

        user_id = get_user_id(conn, email)

        if not user_id:
            raise HTTPException(
              status_code=404,
              detail="User not found"
            )
        conn.execute(
            text("""
                INSERT INTO finance_items (user_id, type, name, amount)
                VALUES (:user_id, :type, :name, :amount)
            """),
            {
                "user_id": user_id,
                "type": data.get("type"),
                "name": get_item_name(data),
                "amount": data.get("amount", 0),
            }
        )

        award_xp(conn, user_id, email, f"finance_{data.get('type')}_created", 30)
        invalidate_finance_caches(email, user_id)
  
    return {"status": "created"}


# =========================
# GET FINANCE
# =========================
@router.get("")
@router.get("/")
def get_finance(user=Depends(get_current_user)):

    email = user

    with engine.connect() as conn:

        user_id = get_user_id(conn, email)

        if not user_id:
            return {"error": "User not found"}

        rows = conn.execute(
            text("""
                SELECT id, type, name, amount
                FROM finance_items
                WHERE user_id = :user_id
                ORDER BY id DESC
            """),
            {"user_id": user_id}
        ).fetchall()

    revenues = []
    charges = []
    debts = []
    savings = []

    for r in rows:

        item = {
            "id": r.id,
            "type": r.type,
            "name": r.name,
            "label": r.name,
            "amount": float(r.amount or 0)
        }

        if r.type == "revenus":
            revenues.append(item)

        elif r.type == "charges":
            charges.append(item)

        elif r.type == "dettes":
            debts.append(item)

        elif r.type == "epargne":
            savings.append(item)

    return {
        "revenus": revenues,
        "charges": charges,
        "dettes": debts,
        "epargne": savings
    }


@router.get("/overview")
def get_finance_overview(user=Depends(get_current_user)):

    email = user

    with engine.connect() as conn:

        user_id = get_user_id(conn, email)

        if not user_id:
            return {"error": "User not found"}

        rows = conn.execute(
            text("""
                SELECT type, COALESCE(SUM(amount), 0) AS total, COUNT(*) AS count
                FROM finance_items
                WHERE user_id = :user_id
                GROUP BY type
            """),
            {"user_id": user_id}
        ).fetchall()

    totals = {
        "revenus": 0.0,
        "charges": 0.0,
        "epargne": 0.0,
        "dettes": 0.0,
    }
    counts = {
        "revenus": 0,
        "charges": 0,
        "epargne": 0,
        "dettes": 0,
    }

    for row in rows:
        if row.type in totals:
            totals[row.type] = float(row.total or 0)
            counts[row.type] = int(row.count or 0)

    income = totals["revenus"]
    expenses = totals["charges"]
    savings = totals["epargne"]
    debt = totals["dettes"]
    cashflow = income - expenses
    savings_rate = (cashflow / income * 100) if income > 0 else 0
    debt_to_income = (debt / income) if income > 0 else 0
    liquid_months = (savings / expenses) if expenses > 0 else 0

    if income <= 0:
        reading = "Ajoute tes revenus suivis pour obtenir une lecture fiable de ta marge de liberte mensuelle."
        priority = "Renseigner les revenus recurrents"
    elif cashflow > 0:
        reading = "Ta base financiere montre une marge mensuelle positive a partir des lignes deja renseignees."
        priority = "Transformer cette marge en coussin de securite et trajectoire d'epargne."
    else:
        reading = "Les charges suivies absorbent les revenus renseignes. La priorite est de retrouver une marge positive."
        priority = "Identifier une charge ajustable ou un revenu complementaire."

    return {
        "version": "finance-overview-v1",
        "source": "finance_items",
        "totals": {
            "income": round(income, 2),
            "expenses": round(expenses, 2),
            "cashflow": round(cashflow, 2),
            "living_margin": round(cashflow, 2),
            "savings": round(savings, 2),
            "debt": round(debt, 2),
        },
        "ratios": {
            "savings_rate": round(savings_rate, 2),
            "debt_to_income": round(debt_to_income, 2),
            "liquid_months": round(liquid_months, 2),
        },
        "counts": counts,
        "reading": reading,
        "priority": priority,
    }


@router.post("/import")
async def import_finance_items(
    file: UploadFile = File(...),
    scope: str = Form("cashflow"),
    user=Depends(get_current_user),
):
    normalized_scope = str(scope or "cashflow").strip().lower()
    if normalized_scope not in {"cashflow", "balance"}:
        raise HTTPException(status_code=400, detail="Scope d'import invalide")

    filename = (file.filename or "").lower()
    if filename.endswith(".pdf") or file.content_type == "application/pdf":
        raise HTTPException(
            status_code=422,
            detail="Import PDF non active sans parseur documentaire. Utilise un CSV ou passe par la Document Inbox.",
        )

    raw = await file.read()
    rows, import_format = read_import_rows(filename, file.content_type, raw)

    inserted = 0
    skipped = 0

    with engine.begin() as conn:
        user_id = get_user_id(conn, user)
        if not user_id:
            raise HTTPException(status_code=404, detail="User not found")

        for row in rows:
            item_type = normalize_finance_type(
                row.get("type") or row.get("category") or row.get("categorie"),
                normalized_scope,
            )
            name = (
                row.get("name")
                or row.get("label")
                or row.get("libelle")
                or row.get("description")
                or ""
            ).strip()
            amount = parse_amount(row.get("amount") or row.get("montant") or row.get("value"))

            if not item_type or not name or amount is None:
                skipped += 1
                continue

            conn.execute(
                text("""
                    INSERT INTO finance_items (user_id, type, name, amount)
                    VALUES (:user_id, :type, :name, :amount)
                """),
                {
                    "user_id": user_id,
                    "type": item_type,
                    "name": name,
                    "amount": amount,
                },
            )
            inserted += 1

        if inserted:
            award_xp(conn, user_id, user, f"finance_{normalized_scope}_imported", 40)
            invalidate_finance_caches(user, user_id)

    return {
        "status": "imported",
        "scope": normalized_scope,
        "format": import_format,
        "inserted": inserted,
        "skipped": skipped,
        "accepted_columns": ["type", "name", "amount"],
    }


# =========================
# UPDATE ITEM
# =========================
@router.put("/{item_id}")
def update_finance(item_id: int, data: dict, user=Depends(get_current_user)):

    email = user

    with engine.begin() as conn:

        user_id = get_user_id(conn, email)

        if not user_id:
            return {"error": "User not found"}

        conn.execute(
            text("""
                UPDATE finance_items
                SET name = :name,
                    amount = :amount
                WHERE id = :id AND user_id = :user_id
            """),
            {
                "id": item_id,
                "user_id": user_id,
                "name": get_item_name(data),
                "amount": data.get("amount", 0)
            }
        )

        award_xp(conn, user_id, email, "finance_updated", 10)
        invalidate_finance_caches(email, user_id)

    return {"status": "updated"}


# =========================
# DELETE ITEM
# =========================
@router.delete("/{item_id}")
def delete_finance(item_id: int, user=Depends(get_current_user)):

    email = user

    with engine.begin() as conn:

        user_id = get_user_id(conn, email)

        if not user_id:
            return {"error": "User not found"}

        conn.execute(
            text("""
                DELETE FROM finance_items
                WHERE id = :id AND user_id = :user_id
            """),
            {
                "id": item_id,
                "user_id": user_id
            }
        )

        invalidate_finance_caches(email, user_id)
        
    return {"status": "deleted"}


# =========================
# ADD XP
# =========================
def add_xp(conn, user_id: int, xp_amount: int):

    row = conn.execute(
        text("""
            SELECT xp
            FROM user_gamification
            WHERE user_id = :user_id
        """),
        {"user_id": user_id}
    ).fetchone()

    if not row:

        conn.execute(
            text("""
                INSERT INTO user_gamification
                (user_id, xp, level)
                VALUES (:user_id, :xp, :level)
            """),
            {
                "user_id": user_id,
                "xp": xp_amount,
                "level": 1
            }
        )

    else:

        new_xp = row.xp + xp_amount
        level = int(new_xp / 100) + 1

        conn.execute(
            text("""
                UPDATE user_gamification
                SET xp = :xp,
                    level = :level
                WHERE user_id = :user_id
            """),
            {
                "xp": new_xp,
                "level": level,
                "user_id": user_id
            }
        )
