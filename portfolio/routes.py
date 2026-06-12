# =========================
# IMPORTS
# =========================
import csv
import io

from core.limiter import limiter
from core.utils import safe_execute
from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from openpyxl import load_workbook
from .schemas import PortfolioRequest
from sqlalchemy import text
from database import engine
from intelligence.gamification.progress_service import award_xp
from workspaces.routes import resolve_workspace_context

from .service import (
    ensure_portfolio_schema,
    get_user_portfolio,
    invalidate_portfolio_cache,
    normalize_asset_type,
    parse_forex_pair,
    save_portfolio_snapshot,
)

router = APIRouter()


def get_user_or_404(conn, email: str):
    user = conn.execute(text("""
        SELECT id FROM users WHERE email = :email
    """), {"email": email}).fetchone()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    return user


def refresh_portfolio_side_effects(user_id: int, email: str):
    invalidate_portfolio_cache(user_id, email)
    save_portfolio_snapshot(user_id)


def build_portfolio_db_payload(data: PortfolioRequest):
    category = normalize_asset_type(data.asset_type)
    asset_name = data.asset_name.upper().strip()
    pair = parse_forex_pair(asset_name) if category == "FOREX" else None

    if category == "FOREX" and not pair:
        raise HTTPException(
            status_code=400,
            detail="Paire FOREX invalide. Exemple attendu: EUR/USD",
        )

    return {
        "asset_name": pair["pair_name"] if pair else asset_name,
        "category": category,
        "quantity": data.quantity,
        "purchase_price": data.purchase_price,
        "pair_name": pair["pair_name"] if pair else None,
        "currency_base": pair["currency_base"] if pair else None,
        "currency_quote": pair["currency_quote"] if pair else None,
    }


def parse_import_float(value):
    cleaned = str(value or "0").strip().replace("\u202f", "").replace(" ", "")
    cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def normalize_import_row(row: dict):
    return {
        str(key or "").strip().lower(): str(value or "").strip()
        for key, value in row.items()
    }


def read_csv_rows(raw: bytes):
    try:
        content = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        content = raw.decode("latin-1")

    reader = csv.DictReader(io.StringIO(content))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV vide ou colonnes manquantes")

    return [normalize_import_row(row) for row in reader]


def read_excel_rows(raw: bytes):
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Fichier Excel invalide ou illisible")

    rows = workbook.active.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        raise HTTPException(status_code=400, detail="Excel vide ou colonnes manquantes")

    normalized_headers = [str(header or "").strip().lower() for header in headers]
    if not any(normalized_headers):
        raise HTTPException(status_code=400, detail="Excel vide ou colonnes manquantes")

    return [
        normalize_import_row(dict(zip(normalized_headers, values)))
        for values in rows
        if any(value not in (None, "") for value in values)
    ]


def read_import_rows(filename: str, content_type: str | None, raw: bytes):
    is_excel = filename.endswith(".xlsx") or content_type == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if is_excel:
        return read_excel_rows(raw), "excel"

    if filename.endswith(".csv") or content_type in {
        "text/csv",
        "application/vnd.ms-excel",
        "application/csv",
    }:
        return read_csv_rows(raw), "csv"

    raise HTTPException(status_code=400, detail="Format supporte: CSV ou Excel (.xlsx)")

# =========================
# GET PORTFOLIO
# =========================
@router.get("")
@router.get("/")
@limiter.limit("10/minute")
def get_portfolio(request: Request):

    def _get():
        user_email = request.state.user_email

        with engine.begin() as conn:

            workspace = resolve_workspace_context(conn, request, user_email)

            return get_user_portfolio(workspace["user_id"], use_cache=False)

    return safe_execute(_get, module_name="PORTFOLIO")


# =========================
# ADD ASSET
# =========================
@router.post("")
@router.post("/")
@limiter.limit("10/minute")
def add_asset(request: Request, data: PortfolioRequest):

    def _add():
        user_email = request.state.user_email

        with engine.begin() as conn:

            workspace = resolve_workspace_context(conn, request, user_email, write=True)
            ensure_portfolio_schema(conn)
            payload = build_portfolio_db_payload(data)

            conn.execute(text("""
                INSERT INTO portfolio (
                    user_id,
                    asset_name,
                    category,
                    quantity,
                    purchase_price,
                    pair_name,
                    currency_base,
                    currency_quote
                )
                VALUES (
                    :user_id,
                    :asset_name,
                    :category,
                    :quantity,
                    :purchase_price,
                    :pair_name,
                    :currency_base,
                    :currency_quote
                )
            """), {
                "user_id": workspace["user_id"],
                **payload,
            })
            
            user_id = workspace["user_id"]
            cache_email = workspace["email"]
            award_xp(conn, user_id, cache_email, "portfolio_asset_created", 50)

        refresh_portfolio_side_effects(user_id, cache_email)
        
        return {"status": "asset ajouté"}

    return safe_execute(_add, module_name="PORTFOLIO")


@router.post("/import")
@limiter.limit("5/minute")
async def import_portfolio_assets(
    request: Request,
    file: UploadFile = File(...),
    scope: str = Form("investments"),
):
    def validate_file():
        filename = (file.filename or "").lower()
        if filename.endswith(".pdf") or file.content_type == "application/pdf":
            raise HTTPException(
                status_code=422,
                detail="Import PDF non active sans parseur documentaire. Utilise un CSV ou passe par la Document Inbox.",
            )

    validate_file()
    filename = (file.filename or "").lower()
    raw = await file.read()
    rows, import_format = read_import_rows(filename, file.content_type, raw)

    def _import():
        user_email = request.state.user_email
        inserted = 0
        skipped = 0

        with engine.begin() as conn:
            workspace = resolve_workspace_context(conn, request, user_email, write=True)
            ensure_portfolio_schema(conn)

            for row in rows:
                asset_name = (
                    row.get("asset_name")
                    or row.get("name")
                    or row.get("ticker")
                    or row.get("symbol")
                    or ""
                ).strip()
                asset_type = (
                    row.get("asset_type")
                    or row.get("type")
                    or row.get("category")
                    or row.get("categorie")
                    or ""
                ).strip()
                quantity = parse_import_float(row.get("quantity") or row.get("quantite"))
                purchase_price = parse_import_float(
                    row.get("purchase_price")
                    or row.get("price")
                    or row.get("prix")
                    or row.get("cost")
                )

                if not asset_name or not asset_type or not quantity or not purchase_price:
                    skipped += 1
                    continue

                try:
                    payload = build_portfolio_db_payload(
                        PortfolioRequest(
                            asset_name=asset_name,
                            asset_type=asset_type,
                            quantity=quantity,
                            purchase_price=purchase_price,
                        )
                    )
                except Exception:
                    skipped += 1
                    continue

                conn.execute(text("""
                    INSERT INTO portfolio (
                        user_id,
                        asset_name,
                        category,
                        quantity,
                        purchase_price,
                        pair_name,
                        currency_base,
                        currency_quote
                    )
                    VALUES (
                        :user_id,
                        :asset_name,
                        :category,
                        :quantity,
                        :purchase_price,
                        :pair_name,
                        :currency_base,
                        :currency_quote
                    )
                """), {
                    "user_id": workspace["user_id"],
                    **payload,
                })
                inserted += 1

            user_id = workspace["user_id"]
            cache_email = workspace["email"]
            if inserted:
                award_xp(conn, user_id, cache_email, "portfolio_imported", 50)

        if inserted:
            refresh_portfolio_side_effects(user_id, cache_email)

        return {
            "status": "imported",
            "scope": scope,
            "format": import_format,
            "inserted": inserted,
            "skipped": skipped,
            "accepted_columns": ["asset_name", "asset_type", "quantity", "purchase_price"],
        }

    return safe_execute(_import, module_name="PORTFOLIO")


# =========================
# UPDATE ASSET
# =========================
@router.put("/{asset_id}")
@limiter.limit("10/minute")
def update_asset(request: Request, asset_id: int, data: PortfolioRequest):

    def _update():
        user_email = request.state.user_email

        with engine.begin() as conn:

            workspace = resolve_workspace_context(conn, request, user_email, write=True)
            ensure_portfolio_schema(conn)
            payload = build_portfolio_db_payload(data)

            result = conn.execute(text("""
                UPDATE portfolio
                SET
                    asset_name = :asset_name,
                    category = :category,
                    quantity = :quantity,
                    purchase_price = :purchase_price,
                    pair_name = :pair_name,
                    currency_base = :currency_base,
                    currency_quote = :currency_quote
                WHERE id = :asset_id
                AND user_id = :user_id
            """), {
                "asset_id": asset_id,
                "user_id": workspace["user_id"],
                **payload,
            })

            if result.rowcount == 0:
                raise HTTPException(status_code=404, detail="Asset not found")

            user_id = workspace["user_id"]
            cache_email = workspace["email"]
            award_xp(conn, user_id, cache_email, "portfolio_asset_updated", 15)

        refresh_portfolio_side_effects(user_id, cache_email)

        return {"status": "updated", "id": asset_id}

    return safe_execute(_update, module_name="PORTFOLIO")


# =========================
# DELETE ASSET
# =========================
@router.delete("/{asset_id}")
@limiter.limit("10/minute")
def delete_asset(request: Request, asset_id: int):

    def _delete():
        user_email = request.state.user_email

        with engine.begin() as conn:

            workspace = resolve_workspace_context(conn, request, user_email, write=True)

            result = conn.execute(text("""
                DELETE FROM portfolio
                WHERE id = :asset_id AND user_id = :user_id
            """), {
                "asset_id": asset_id,
                "user_id": workspace["user_id"]
            })

            if result.rowcount == 0:
                raise HTTPException(
                    status_code=404,
                    detail="Asset not found or not owned by user",
                )

            user_id = workspace["user_id"]
            cache_email = workspace["email"]

        refresh_portfolio_side_effects(user_id, cache_email)

        return {"status": "deleted", "id": asset_id}

    return safe_execute(_delete, module_name="PORTFOLIO")


# =========================
# PORTFOLIO HISTORY
# =========================
@router.get("/history")
def portfolio_history(request: Request):

    user_email = request.state.user_email

    with engine.begin() as conn:
        workspace = resolve_workspace_context(conn, request, user_email)

        total_cost = conn.execute(text("""
            SELECT COALESCE(SUM(quantity * purchase_price), 0)
            FROM portfolio
            WHERE user_id = :user_id
        """), {"user_id": workspace["user_id"]}).scalar()

        rows = conn.execute(text("""
            SELECT total_value, created_at
            FROM portfolio_history
            WHERE user_id = :user_id
            ORDER BY created_at ASC
        """), {"user_id": workspace["user_id"]}).fetchall()

    return {
        "history": [
            {
                "date": r.created_at.isoformat(),
                "value": float(r.total_value),
                "cost": float(total_cost or 0),
                "gain": float(r.total_value) - float(total_cost or 0)
            }
            for r in rows
        ]
    }
